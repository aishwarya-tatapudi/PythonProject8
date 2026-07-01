import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.svm import SVC
from sklearn.metrics import (
    accuracy_score,
    confusion_matrix,
    classification_report
)
from sentence_transformers import SentenceTransformer
import warnings
warnings.filterwarnings('ignore')

# ============================================
# STEP 1: LOAD DATA
# ============================================
print("="*60)
print("STEP 1: LOADING DATA")
print("="*60)

# Load dataset - handles both .xls and .xlsx
try:
    df = pd.read_excel(
        'PromptInjectionPrompts.xlsx',
        engine='openpyxl'
    )
except FileNotFoundError:
    df = pd.read_excel(
        'PromptInjectionPrompts.xls',
        engine='xlrd'
    )

# Show actual column names to catch schema mismatches early
print(f"Columns found: {df.columns.tolist()}")

# Rename columns to standard names
df = df.rename(columns={
    'Prompts': 'text',
    'Label': 'label',
    'Language': 'language'
})

# Verify required columns exist after rename
required_cols = ['text', 'label', 'language']
missing = [c for c in required_cols if c not in df.columns]
if missing:
    raise ValueError(
        f"Missing columns after rename: {missing}. "
        f"Check that your Excel headers match: 'Prompts', 'Label', 'Language'"
    )

# Keep only what we need
df = df[required_cols]

print(f"Total examples: {len(df)}")
print(f"\nLanguage distribution:")
print(df['language'].value_counts())
print(f"\nLabel distribution:")
print(df['label'].value_counts())

# Safe sample display — won't crash if subset is empty
hindi_attacks = df[(df['language'] == 'Hindi') & (df['label'] == 1)]['text']
hindi_innocent = df[(df['language'] == 'Hindi') & (df['label'] == 0)]['text']

if not hindi_attacks.empty:
    print(f"\nSample Hindi attack:")
    print(hindi_attacks.iloc[0])
else:
    print("\nNo Hindi attack samples found.")

if not hindi_innocent.empty:
    print(f"\nSample Hindi innocent:")
    print(hindi_innocent.iloc[0])
else:
    print("\nNo Hindi innocent samples found.")

# ============================================
# STEP 2: SPLIT DATA
# ============================================
print("\n" + "="*60)
print("STEP 2: SPLITTING DATA 80/20")
print("="*60)

# Separate by language
hindi_df = df[df['language'] == 'Hindi'].copy()
hinglish_df = df[df['language'] == 'Hinglish'].copy()

print(f"Hindi total: {len(hindi_df)}")
print(f"Hinglish total: {len(hinglish_df)}")

# Split each language separately 80/20
hi_train, hi_test = train_test_split(
    hindi_df,
    test_size=0.2,
    random_state=42,
    stratify=hindi_df['label']
)

hg_train, hg_test = train_test_split(
    hinglish_df,
    test_size=0.2,
    random_state=42,
    stratify=hinglish_df['label']
)

# Combine training data from both languages
train_df = pd.concat([hi_train, hg_train], ignore_index=True)

print(f"\nTraining set: {len(train_df)}")
print(f"  Hindi train: {len(hi_train)}")
print(f"  Hinglish train: {len(hg_train)}")
print(f"\nTest sets:")
print(f"  Hindi test: {len(hi_test)}")
print(f"  Hinglish test: {len(hg_test)}")

# ============================================
# STEP 3: LOAD SENTENCE TRANSFORMER
# ============================================
print("\n" + "="*60)
print("STEP 3: LOADING MULTILINGUAL MODEL")
print("="*60)

print("Loading paraphrase-multilingual-MiniLM-L12-v2...")
print("(This understands Hindi and Hinglish meaning)")
embedder = SentenceTransformer(
    'paraphrase-multilingual-MiniLM-L12-v2'
)
print("Model loaded!")

# ============================================
# STEP 4: CONVERT TEXT TO MEANING VECTORS
# ============================================
print("\n" + "="*60)
print("STEP 4: CONVERTING TEXT TO MEANING VECTORS")
print("="*60)

print("Converting training data...")
train_vectors = embedder.encode(
    train_df['text'].tolist(),
    show_progress_bar=True,
    batch_size=32
)
print(f"Training vectors shape: {train_vectors.shape}")

print("\nConverting Hindi test data...")
hi_test_vectors = embedder.encode(
    hi_test['text'].tolist(),
    show_progress_bar=True,
    batch_size=32
)

print("\nConverting Hinglish test data...")
hg_test_vectors = embedder.encode(
    hg_test['text'].tolist(),
    show_progress_bar=True,
    batch_size=32
)

# ============================================
# STEP 5: TRAIN SVM
# ============================================
print("\n" + "="*60)
print("STEP 5: TRAINING SVM CLASSIFIER")
print("="*60)

print("Training SVM on Hindi + Hinglish data...")
svm = SVC(
    kernel='rbf',
    probability=True,
    class_weight='balanced',
    random_state=42,
    C=1.0
)

svm.fit(train_vectors, train_df['label'].values)
print("SVM training complete!")

# ============================================
# STEP 6: EVALUATE RESULTS
# ============================================
print("\n" + "="*60)
print("STEP 6: EVALUATING RESULTS")
print("="*60)

def evaluate(name, vectors, true_labels):
    """
    Evaluate model performance for one language.
    Returns accuracy, FPR, FNR, and raw confusion matrix counts.
    """
    predictions = svm.predict(vectors)

    tn, fp, fn, tp = confusion_matrix(
        true_labels,
        predictions
    ).ravel()

    accuracy = (tp + tn) / (tp + tn + fp + fn)
    fpr = fp / (fp + tn) if (fp + tn) > 0 else 0        # innocent wrongly blocked
    fnr = fn / (fn + tp) if (fn + tp) > 0 else 0        # attacks that slipped through
    detection_rate = tp / (tp + fn) if (tp + fn) > 0 else 0

    print(f"\n{'='*40}")
    print(f"RESULTS: {name}")
    print(f"{'='*40}")
    print(f"Accuracy:                {accuracy:.2%}")
    print(f"Detection Rate (TPR):    {detection_rate:.2%}")
    print(f"False Positive Rate:     {fpr:.2%}")
    print(f"  (innocent users wrongly blocked)")
    print(f"False Negative Rate:     {fnr:.2%}")
    print(f"  (attacks that slipped through)")
    print(f"\nConfusion Matrix:")
    print(f"  True Positives:  {tp} (attacks correctly caught)")
    print(f"  True Negatives:  {tn} (innocent correctly allowed)")
    print(f"  False Positives: {fp} (innocent wrongly blocked)")
    print(f"  False Negatives: {fn} (attacks missed)")

    return {
        'language': name,
        'accuracy': accuracy,
        'detection_rate': detection_rate,
        'fpr': fpr,
        'fnr': fnr,
        'tp': int(tp),
        'tn': int(tn),
        'fp': int(fp),
        'fn': int(fn),
        'n_samples': len(true_labels)
    }

# Evaluate both languages
hindi_results = evaluate("HINDI", hi_test_vectors, hi_test['label'].values)
hinglish_results = evaluate("HINGLISH", hg_test_vectors, hg_test['label'].values)

# ============================================
# STEP 7: COMPARISON SUMMARY
# ============================================
print("\n" + "="*60)
print("STEP 7: FINAL COMPARISON")
print("="*60)

print(f"\n{'Metric':<30}{'Hindi':>12}{'Hinglish':>12}")
print("-"*55)
print(f"{'Accuracy':<30}{hindi_results['accuracy']:>12.2%}{hinglish_results['accuracy']:>12.2%}")
print(f"{'Detection Rate':<30}{hindi_results['detection_rate']:>12.2%}{hinglish_results['detection_rate']:>12.2%}")
print(f"{'False Positive Rate':<30}{hindi_results['fpr']:>12.2%}{hinglish_results['fpr']:>12.2%}")
print(f"{'False Negative Rate':<30}{hindi_results['fnr']:>12.2%}{hinglish_results['fnr']:>12.2%}")

# Fairness gap
fpr_gap = abs(hindi_results['fpr'] - hinglish_results['fpr'])
fnr_gap = abs(hindi_results['fnr'] - hinglish_results['fnr'])

print(f"\nFAIRNESS ANALYSIS:")
print(f"FPR gap between languages: {fpr_gap:.2%}")
print(f"FNR gap between languages: {fnr_gap:.2%}")

if hindi_results['fpr'] > hinglish_results['fpr']:
    print(f"→ Hindi users wrongly blocked MORE than Hinglish")
elif hinglish_results['fpr'] > hindi_results['fpr']:
    print(f"→ Hinglish users wrongly blocked MORE than Hindi")
else:
    print(f"→ Equal false positive rates — no fairness gap")

# Weighted combined accuracy (accounts for unequal test set sizes)
total_samples = hindi_results['n_samples'] + hinglish_results['n_samples']
combined_acc = (
    hindi_results['accuracy'] * hindi_results['n_samples'] +
    hinglish_results['accuracy'] * hinglish_results['n_samples']
) / total_samples

print(f"\nCOMPARISON WITH SRINIVASAN ET AL. (2026):")
print(f"Their accuracy:    99.70%")
print(f"Your accuracy:     {combined_acc:.2%}")
print(f"Difference:        {abs(99.70 - combined_acc * 100):.2f}%")

# Save results
results_df = pd.DataFrame([hindi_results, hinglish_results])
results_df.to_csv('results_model1.csv', index=False)
print(f"\nResults saved to results_model1.csv")
attack_confidence = svm.predict_proba(hg_test_vectors)[:, 1]

critical = np.sum(attack_confidence >= 0.90)
high     = np.sum((attack_confidence >= 0.70) & (attack_confidence < 0.90))
medium   = np.sum((attack_confidence >= 0.50) & (attack_confidence < 0.70))
low      = np.sum((attack_confidence >= 0.30) & (attack_confidence < 0.50))
safe     = np.sum(attack_confidence < 0.30)

print(f"Critical (auto block):     {critical} ({critical/400:.1%})")
print(f"High (urgent review):      {high}     ({high/400:.1%})")
print(f"Medium (human review):     {medium}   ({medium/400:.1%})")
print(f"Low (log and monitor):     {low}      ({low/400:.1%})")
print(f"Safe (auto allow):         {safe}     ({safe/400:.1%})")

# ============================================
# ADD THIS AFTER YOUR EXISTING STEP 7
# Requires: svm, hi_test, hg_test,
#           hi_test_vectors, hg_test_vectors
# ============================================

# ============================================
# STEP 8: FIVE-TIER CONFIDENCE SCORING
# ============================================
print("\n" + "="*60)
print("STEP 8: FIVE-TIER CONFIDENCE RISK SCORING")
print("="*60)

def assign_tier(conf):
    """
    Maps attack confidence score to risk tier.
    Confidence = closeness to attack boundary (1.0 = certain attack).
    """
    if conf >= 0.90:
        return 'CRITICAL'   # auto block
    elif conf >= 0.70:
        return 'HIGH'       # urgent CSR review same day
    elif conf >= 0.50:
        return 'MEDIUM'     # standard CSR review next day
    elif conf >= 0.30:
        return 'LOW'        # log and monitor
    else:
        return 'SAFE'       # auto allow

def tier_distribution(name, vectors, n_samples):
    confidences = svm.predict_proba(vectors)[:, 1]
    tiers       = np.array([assign_tier(c) for c in confidences])

    critical = np.sum(tiers == 'CRITICAL')
    high     = np.sum(tiers == 'HIGH')
    medium   = np.sum(tiers == 'MEDIUM')
    low      = np.sum(tiers == 'LOW')
    safe     = np.sum(tiers == 'SAFE')

    print(f"\n{name} Tier Distribution ({n_samples} prompts):")
    print(f"  Critical (auto block):    {critical:>4} ({critical/n_samples:.1%}) → immediate block")
    print(f"  High     (urgent CSR):    {high:>4} ({high/n_samples:.1%}) → same day review")
    print(f"  Medium   (standard CSR):  {medium:>4} ({medium/n_samples:.1%}) → next day review")
    print(f"  Low      (log/monitor):   {low:>4} ({low/n_samples:.1%}) → batch review")
    print(f"  Safe     (auto allow):    {safe:>4} ({safe/n_samples:.1%}) → no action")
    print(f"\n  Auto-handled: {critical+safe} ({(critical+safe)/n_samples:.1%}) — no human needed")
    print(f"  Needs CSR:    {high+medium} ({(high+medium)/n_samples:.1%}) — human review required")

    return confidences, tiers

hi_confidences, hi_tiers = tier_distribution("HINDI",    hi_test_vectors, len(hi_test))
hg_confidences, hg_tiers = tier_distribution("HINGLISH", hg_test_vectors, len(hg_test))

# ============================================
# STEP 9: PROMPT INSPECTION
# ============================================
print("\n" + "="*60)
print("STEP 9: PROMPT INSPECTION — FLAGGED PROMPTS")
print("="*60)

def inspect_prompts(name, test_df, confidences, tiers):
    # reset_index prevents pandas/numpy index mismatch
    inspect_df = test_df.copy().reset_index(drop=True)
    inspect_df['attack_confidence'] = confidences
    inspect_df['tier']              = tiers
    inspect_df['true_label']        = inspect_df['label']

    # --- FALSE POSITIVES ---
    # Innocent prompts (label=0) wrongly flagged as CRITICAL or HIGH
    fp_df = inspect_df[
        (inspect_df['true_label'] == 0) &
        (inspect_df['tier'].isin(['CRITICAL', 'HIGH']))
    ].sort_values('attack_confidence', ascending=False)

    print(f"\n{'='*50}")
    print(f"FALSE POSITIVES — {name} ({len(fp_df)} prompts)")
    print(f"Innocent prompts wrongly blocked")
    print(f"{'='*50}")
    for _, row in fp_df.iterrows():
        print(f"Confidence: {row['attack_confidence']:.2%} | Tier: {row['tier']}")
        print(f"Prompt: {row['text']}")
        print("-"*40)

    # --- FALSE NEGATIVES ---
    # Real attacks (label=1) that scored LOW or SAFE
    fn_df = inspect_df[
        (inspect_df['true_label'] == 1) &
        (inspect_df['tier'].isin(['LOW', 'SAFE']))
    ].sort_values('attack_confidence', ascending=True)

    print(f"\n{'='*50}")
    print(f"FALSE NEGATIVES — {name} ({len(fn_df)} prompts)")
    print(f"Real attacks that slipped through")
    print(f"{'='*50}")
    for _, row in fn_df.iterrows():
        print(f"Confidence: {row['attack_confidence']:.2%} | Tier: {row['tier']}")
        print(f"Prompt: {row['text']}")
        print("-"*40)

    # --- PROMPTS BY TIER ---
    for tier in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW', 'SAFE']:
        tier_df = inspect_df[
            inspect_df['tier'] == tier
        ].sort_values('attack_confidence', ascending=False)

        print(f"\n{'='*50}")
        print(f"{tier} TIER — {name} ({len(tier_df)} prompts)")
        print(f"{'='*50}")
        for _, row in tier_df.head(5).iterrows():  # show top 5 per tier
            actual = "ATTACK" if row['true_label'] == 1 else "INNOCENT"
            correct = "✓" if (
                (row['true_label'] == 1 and row['tier'] in ['CRITICAL', 'HIGH']) or
                (row['true_label'] == 0 and row['tier'] in ['LOW', 'SAFE'])
            ) else "✗ WRONG"
            print(f"Confidence: {row['attack_confidence']:.2%} | Actually: {actual} {correct}")
            print(f"Prompt: {row['text']}")
            print("-"*40)

    # save full results
    filename = f"{name.lower()}_inspection.csv"
    inspect_df.to_csv(filename, index=False)
    print(f"\nFull {name} inspection saved to {filename}")

    return inspect_df

hi_inspect = inspect_prompts("HINDI",    hi_test, hi_confidences, hi_tiers)
hg_inspect = inspect_prompts("HINGLISH", hg_test, hg_confidences, hg_tiers)

# ============================================
# STEP 10: FAIRNESS GAP SUMMARY
# ============================================
print("\n" + "="*60)
print("STEP 10: FAIRNESS GAP — HINDI VS HINGLISH")
print("="*60)

hi_fp  = len(hi_inspect[(hi_inspect['true_label'] == 0) & (hi_inspect['tier'].isin(['CRITICAL', 'HIGH']))])
hg_fp  = len(hg_inspect[(hg_inspect['true_label'] == 0) & (hg_inspect['tier'].isin(['CRITICAL', 'HIGH']))])
hi_fn  = len(hi_inspect[(hi_inspect['true_label'] == 1) & (hi_inspect['tier'].isin(['LOW', 'SAFE']))])
hg_fn  = len(hg_inspect[(hg_inspect['true_label'] == 1) & (hg_inspect['tier'].isin(['LOW', 'SAFE']))])
hi_csr = len(hi_inspect[hi_inspect['tier'].isin(['HIGH', 'MEDIUM'])])
hg_csr = len(hg_inspect[hg_inspect['tier'].isin(['HIGH', 'MEDIUM'])])

print(f"\n{'Metric':<40}{'Hindi':>10}{'Hinglish':>10}")
print("-"*60)
print(f"{'False Positives (wrongly blocked)':<40}{hi_fp:>10}{hg_fp:>10}")
print(f"{'False Negatives (attacks missed)':<40}{hi_fn:>10}{hg_fn:>10}")
print(f"{'CSR queue (High + Medium)':<40}{hi_csr:>10}{hg_csr:>10}")

print(f"\nKEY FINDINGS:")
if hg_fp > hi_fp:
    print(f"→ Hinglish has {hg_fp - hi_fp} more false positives than Hindi")
    print(f"  Innocent Hinglish users are wrongly blocked more often")
if hg_fn > hi_fn:
    print(f"→ Hinglish has {hg_fn - hi_fn} more attacks slipping through than Hindi")
    print(f"  Hinglish attack patterns are harder for the model to detect")
if hg_csr > hi_csr:
    print(f"→ Hinglish needs {hg_csr - hi_csr} more CSR reviews than Hindi")
    print(f"  Higher operational burden for Hinglish users")
# ============================================
# ADD THIS AT THE END OF YOUR SCRIPT
# Requires: hi_inspect, hg_inspect
# (these are created in Step 9)
# ============================================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ============================================
# STEP 11: EXPORT ALL PROMPTS TO EXCEL
# ============================================
print("\n" + "="*60)
print("STEP 11: EXPORTING PROMPTS TO EXCEL")
print("="*60)

# Color scheme for each tier
TIER_COLORS = {
    'CRITICAL': 'FF4444',   # red
    'HIGH':     'FF8C00',   # orange
    'MEDIUM':   'FFD700',   # yellow
    'LOW':      '90EE90',   # light green
    'SAFE':     '4CAF50',   # green
}

# Color scheme for correct vs wrong
CORRECT_COLOR = 'C8E6C9'    # light green
WRONG_COLOR   = 'FFCDD2'    # light red

def style_header(cell, bg_color='1F4E79'):
    cell.font      = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    cell.fill      = PatternFill('solid', start_color=bg_color)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_tier_cell(cell, tier):
    color = TIER_COLORS.get(tier, 'FFFFFF')
    cell.fill      = PatternFill('solid', start_color=color)
    cell.font      = Font(bold=True, name='Arial', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def style_data_cell(cell, correct=True):
    color = CORRECT_COLOR if correct else WRONG_COLOR
    cell.fill      = PatternFill('solid', start_color=color)
    cell.font      = Font(name='Arial', size=10)
    cell.alignment = Alignment(vertical='center', wrap_text=True)

def write_sheet(ws, df, language):
    """Write one language's prompts to a worksheet organized by tier."""

    # --- Column headers ---
    headers = ['#', 'Prompt', 'True Label', 'Attack Confidence', 'Tier', 'Correct?']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header(cell)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 6    # #
    ws.column_dimensions['B'].width = 70   # Prompt
    ws.column_dimensions['C'].width = 14   # True Label
    ws.column_dimensions['D'].width = 20   # Attack Confidence
    ws.column_dimensions['E'].width = 12   # Tier
    ws.column_dimensions['F'].width = 12   # Correct?

    # --- Freeze header row ---
    ws.freeze_panes = 'A2'

    # --- Write prompts grouped by tier ---
    row = 2
    tier_order = ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW', 'SAFE']

    for tier in tier_order:
        tier_df = df[df['tier'] == tier].sort_values(
            'attack_confidence', ascending=(tier in ['LOW', 'SAFE'])
        )

        if tier_df.empty:
            continue

        # Tier section header
        section_cell = ws.cell(row=row, column=1,
            value=f"{tier} — {len(tier_df)} prompts ({len(tier_df)/len(df):.1%})")
        section_cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
        section_cell.fill = PatternFill('solid', start_color=TIER_COLORS[tier])
        section_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 20
        row += 1

        # Prompts in this tier
        for i, (_, prompt_row) in enumerate(tier_df.iterrows(), 1):
            true_label  = 'ATTACK'   if prompt_row['true_label'] == 1 else 'INNOCENT'
            confidence  = f"{prompt_row['attack_confidence']:.2%}"
            tier_val    = prompt_row['tier']

            # Is the classification correct?
            is_correct = (
                (prompt_row['true_label'] == 1 and tier_val in ['CRITICAL', 'HIGH']) or
                (prompt_row['true_label'] == 0 and tier_val in ['LOW', 'SAFE'])
            )
            correct_str = '✓ Correct' if is_correct else '✗ Wrong'

            # Write row
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=prompt_row['text'])
            ws.cell(row=row, column=3, value=true_label)
            ws.cell(row=row, column=4, value=confidence)
            ws.cell(row=row, column=5, value=tier_val)
            ws.cell(row=row, column=6, value=correct_str)

            # Style each cell
            style_tier_cell(ws.cell(row=row, column=5), tier_val)
            style_data_cell(ws.cell(row=row, column=6), is_correct)

            for col in [1, 2, 3, 4]:
                ws.cell(row=row, column=col).font      = Font(name='Arial', size=10)
                ws.cell(row=row, column=col).alignment = Alignment(
                    vertical='center', wrap_text=True
                )

            ws.row_dimensions[row].height = 30
            row += 1

        row += 1  # blank row between tiers

    print(f"  {language}: {len(df)} prompts written across {len(tier_order)} tiers")

def write_summary_sheet(ws, hi_df, hg_df):
    """Write a summary comparison sheet."""

    # Title
    title = ws.cell(row=1, column=1, value='FAIRNESS ANALYSIS SUMMARY')
    title.font      = Font(bold=True, color='FFFFFF', name='Arial', size=14)
    title.fill      = PatternFill('solid', start_color='1F4E79')
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 30

    # Headers
    headers = ['Metric', 'Hindi', 'Hinglish', 'Gap']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        style_header(cell)

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    # Calculate metrics
    def get_metrics(df):
        total   = len(df)
        fp      = len(df[(df['true_label'] == 0) & (df['tier'].isin(['CRITICAL', 'HIGH']))])
        fn      = len(df[(df['true_label'] == 1) & (df['tier'].isin(['LOW', 'SAFE']))])
        csr     = len(df[df['tier'].isin(['HIGH', 'MEDIUM'])])
        correct = len(df[
            ((df['true_label'] == 1) & (df['tier'].isin(['CRITICAL', 'HIGH']))) |
            ((df['true_label'] == 0) & (df['tier'].isin(['LOW', 'SAFE'])))
        ])
        accuracy = correct / total
        fpr      = fp / len(df[df['true_label'] == 0]) if len(df[df['true_label'] == 0]) > 0 else 0
        fnr      = fn / len(df[df['true_label'] == 1]) if len(df[df['true_label'] == 1]) > 0 else 0

        tier_counts = {t: len(df[df['tier'] == t]) for t in ['CRITICAL','HIGH','MEDIUM','LOW','SAFE']}
        return {
            'total': total, 'fp': fp, 'fn': fn, 'csr': csr,
            'accuracy': accuracy, 'fpr': fpr, 'fnr': fnr,
            **tier_counts
        }

    hi  = get_metrics(hi_df)
    hg  = get_metrics(hg_df)

    # Rows
    metrics = [
        ('Total Prompts',              hi['total'],    hg['total'],    None),
        ('Accuracy',                   f"{hi['accuracy']:.2%}", f"{hg['accuracy']:.2%}", None),
        ('False Positive Rate (FPR)',  f"{hi['fpr']:.2%}", f"{hg['fpr']:.2%}", f"{abs(hi['fpr']-hg['fpr']):.2%}"),
        ('False Negative Rate (FNR)',  f"{hi['fnr']:.2%}", f"{hg['fnr']:.2%}", f"{abs(hi['fnr']-hg['fnr']):.2%}"),
        ('False Positives (count)',    hi['fp'],       hg['fp'],       hg['fp'] - hi['fp']),
        ('False Negatives (count)',    hi['fn'],       hg['fn'],       hg['fn'] - hi['fn']),
        ('CSR Queue Size',             hi['csr'],      hg['csr'],      hg['csr'] - hi['csr']),
        ('',                           '',             '',             ''),
        ('CRITICAL tier',              hi['CRITICAL'], hg['CRITICAL'], hg['CRITICAL'] - hi['CRITICAL']),
        ('HIGH tier',                  hi['HIGH'],     hg['HIGH'],     hg['HIGH'] - hi['HIGH']),
        ('MEDIUM tier',                hi['MEDIUM'],   hg['MEDIUM'],   hg['MEDIUM'] - hi['MEDIUM']),
        ('LOW tier',                   hi['LOW'],      hg['LOW'],      hg['LOW'] - hi['LOW']),
        ('SAFE tier',                  hi['SAFE'],     hg['SAFE'],     hg['SAFE'] - hi['SAFE']),
    ]

    for r, (metric, hi_val, hg_val, gap) in enumerate(metrics, 3):
        ws.cell(row=r, column=1, value=metric).font = Font(name='Arial', size=10, bold=(metric in ['', 'CRITICAL tier']))
        ws.cell(row=r, column=2, value=hi_val).alignment  = Alignment(horizontal='center')
        ws.cell(row=r, column=3, value=hg_val).alignment  = Alignment(horizontal='center')
        ws.cell(row=r, column=4, value=gap).alignment     = Alignment(horizontal='center')

        # Highlight gap cells where Hinglish is worse
        if gap and isinstance(gap, (int, float)) and gap > 0:
            ws.cell(row=r, column=4).fill = PatternFill('solid', start_color='FFCDD2')
        elif gap and isinstance(gap, str) and gap != '':
            ws.cell(row=r, column=4).fill = PatternFill('solid', start_color='FFCDD2')

        ws.row_dimensions[r].height = 20

    print(f"  Summary sheet written")

# --- Build workbook ---
wb = Workbook()

# Sheet 1: Summary
ws_summary = wb.active
ws_summary.title = 'Summary'
write_summary_sheet(ws_summary, hi_inspect, hg_inspect)

# Sheet 2: Hindi prompts by tier
ws_hindi = wb.create_sheet('Hindi — All Prompts')
write_sheet(ws_hindi, hi_inspect, 'Hindi')

# Sheet 3: Hinglish prompts by tier
ws_hinglish = wb.create_sheet('Hinglish — All Prompts')
write_sheet(ws_hinglish, hg_inspect, 'Hinglish')

# Sheet 4: False positives only (both languages)
ws_fp = wb.create_sheet('False Positives')
fp_all = []
for lang, df in [('Hindi', hi_inspect), ('Hinglish', hg_inspect)]:
    fp_df = df[(df['true_label'] == 0) & (df['tier'].isin(['CRITICAL', 'HIGH']))].copy()
    fp_df['language'] = lang
    fp_all.append(fp_df)

fp_combined = pd.concat(fp_all, ignore_index=True).sort_values(
    'attack_confidence', ascending=False
)

fp_headers = ['#', 'Language', 'Prompt', 'Attack Confidence', 'Tier']
for col, h in enumerate(fp_headers, 1):
    cell = ws_fp.cell(row=1, column=col, value=h)
    style_header(cell)

ws_fp.column_dimensions['A'].width = 6
ws_fp.column_dimensions['B'].width = 12
ws_fp.column_dimensions['C'].width = 70
ws_fp.column_dimensions['D'].width = 20
ws_fp.column_dimensions['E'].width = 12
ws_fp.freeze_panes = 'A2'

for i, (_, row_data) in enumerate(fp_combined.iterrows(), 1):
    ws_fp.cell(row=i+1, column=1, value=i)
    ws_fp.cell(row=i+1, column=2, value=row_data['language'])
    ws_fp.cell(row=i+1, column=3, value=row_data['text'])
    ws_fp.cell(row=i+1, column=4, value=f"{row_data['attack_confidence']:.2%}")
    cell_tier = ws_fp.cell(row=i+1, column=5, value=row_data['tier'])
    style_tier_cell(cell_tier, row_data['tier'])
    ws_fp.row_dimensions[i+1].height = 30
    for col in [1, 2, 3, 4]:
        ws_fp.cell(row=i+1, column=col).font      = Font(name='Arial', size=10)
        ws_fp.cell(row=i+1, column=col).fill      = PatternFill('solid', start_color='FFCDD2')
        ws_fp.cell(row=i+1, column=col).alignment = Alignment(vertical='center', wrap_text=True)

print(f"  False Positives sheet: {len(fp_combined)} prompts")

# Sheet 5: False negatives only (both languages)
ws_fn = wb.create_sheet('False Negatives')
fn_all = []
for lang, df in [('Hindi', hi_inspect), ('Hinglish', hg_inspect)]:
    fn_df = df[(df['true_label'] == 1) & (df['tier'].isin(['LOW', 'SAFE']))].copy()
    fn_df['language'] = lang
    fn_all.append(fn_df)

fn_combined = pd.concat(fn_all, ignore_index=True).sort_values(
    'attack_confidence', ascending=True
)

fn_headers = ['#', 'Language', 'Prompt', 'Attack Confidence', 'Tier']
for col, h in enumerate(fn_headers, 1):
    cell = ws_fn.cell(row=1, column=col, value=h)
    style_header(cell)

ws_fn.column_dimensions['A'].width = 6
ws_fn.column_dimensions['B'].width = 12
ws_fn.column_dimensions['C'].width = 70
ws_fn.column_dimensions['D'].width = 20
ws_fn.column_dimensions['E'].width = 12
ws_fn.freeze_panes = 'A2'

for i, (_, row_data) in enumerate(fn_combined.iterrows(), 1):
    ws_fn.cell(row=i+1, column=1, value=i)
    ws_fn.cell(row=i+1, column=2, value=row_data['language'])
    ws_fn.cell(row=i+1, column=3, value=row_data['text'])
    ws_fn.cell(row=i+1, column=4, value=f"{row_data['attack_confidence']:.2%}")
    cell_tier = ws_fn.cell(row=i+1, column=5, value=row_data['tier'])
    style_tier_cell(cell_tier, row_data['tier'])
    ws_fn.row_dimensions[i+1].height = 30
    for col in [1, 2, 3, 4]:
        ws_fn.cell(row=i+1, column=col).font      = Font(name='Arial', size=10)
        ws_fn.cell(row=i+1, column=col).fill      = PatternFill('solid', start_color='FFF9C4')
        ws_fn.cell(row=i+1, column=col).alignment = Alignment(vertical='center', wrap_text=True)

print(f"  False Negatives sheet: {len(fn_combined)} prompts")

# --- Save ---
output_path = 'prompt_inspection_results.xlsx'
wb.save(output_path)
print(f"\nExcel file saved: {output_path}")
print("\nSheets created:")
print("  1. Summary            — fairness gap metrics side by side")
print("  2. Hindi — All Prompts  — all Hindi prompts by tier")
print("  3. Hinglish — All Prompts — all Hinglish prompts by tier")
print("  4. False Positives    — innocent prompts wrongly blocked")
print("  5. False Negatives    — real attacks that slipped through")